import pytest
from unittest.mock import MagicMock, patch
from openpyxl import Workbook
from process.excel_process import ExcelProcess


@pytest.fixture
def mock_connections():
    """Simula las conexiones a Oracle y PostgreSQL."""
    oc_connection = MagicMock()
    pg_connection = MagicMock()
    return oc_connection, pg_connection


@pytest.fixture
def mock_workbook():
    """Crea un libro de Excel simulado con pocas filas."""
    wb = Workbook()
    ws = wb.active
    ws.append(["ID"])  # Agregar encabezado
    ws.append([12345])  # Primera fila de datos
    ws.append([67890])  # Segunda fila de datos
    return wb


@pytest.fixture
def excel_instance(mock_connections):
    """Crea una instancia de ExcelProcess para las pruebas."""
    oc_connection, pg_connection = mock_connections
    return ExcelProcess(oc_connection, pg_connection)


def test_execute_process_with_data(excel_instance, mock_workbook):
    """Prueba el proceso de lectura de Excel y la ejecución del método _define_what_process_run."""
    # Simula el método load_workbook para que devuelva el mock_workbook en lugar del archivo real
    with patch("openpyxl.load_workbook", return_value=mock_workbook), \
         patch.object(excel_instance, '_define_what_process_run') as mock_define_process, \
         patch("openpyxl.worksheet.worksheet.Worksheet.iter_rows", return_value=[("ID",), (12345,), (67890,)]):
        

        excel_instance.execute()

        assert mock_define_process.call_count == 2
        mock_define_process.assert_any_call(None, 12345)
        mock_define_process.assert_any_call(None, 67890)


def test_execute_process_without_data(excel_instance):
    """Prueba la ejecución cuando el archivo de Excel no tiene datos (además del encabezado)."""

    empty_workbook = Workbook()
    empty_workbook.active.append(["ID"])  

    with patch("openpyxl.load_workbook", return_value=empty_workbook), \
         patch.object(excel_instance, '_define_what_process_run') as mock_define_process, \
         patch("openpyxl.worksheet.worksheet.Worksheet.iter_rows", return_value=[("ID",)]):
        
        excel_instance.execute()

        mock_define_process.assert_not_called()

    
@patch('process.excel_process.load_workbook')  
def test_execute_without_data_no_exception(mock_load_workbook, excel_instance):
    """Prueba la ejecución cuando no hay datos (solo encabezado) sin excepciones."""
    empty_workbook = Workbook()

    ws = empty_workbook.active
    ws.append(["ID"])  

    mock_load_workbook.return_value = empty_workbook

    with patch("process.excel_process.ExcelProcess._define_what_process_run") as mock_define_process:
        excel_instance.execute()


        mock_define_process.assert_not_called()
        

def test_define_what_process_run_imports_correct_class(excel_instance, mock_connections):
    """Prueba que se importe la clase correcta desde el módulo correspondiente."""
    
    record = {"another_field": "another_value"}
    fic_id = 456

    mock_class = MagicMock()

    mock_module = MagicMock()
    mock_module.FichaCaracterizacion = mock_class  

    with patch('importlib.import_module', return_value=mock_module):
        excel_instance._define_what_process_run(record, fic_id)

        mock_class.assert_called_once_with(
            excel_instance.oc_connection,
            excel_instance.pg_connection,
            record,
            fic_id
        )

        instance_class = mock_class.return_value  
        instance_class.process.assert_called_once()  

def test_define_what_process_run_calls_process(excel_instance, mock_connections):
    """Prueba que el método process se llame correctamente en la instancia de clase importada."""
    
    record = {"some_field": "some_value"}
    fic_id = 123

    mock_class = MagicMock()
    mock_class_instance = mock_class.return_value  

    mock_module = MagicMock()
    mock_module.FichaCaracterizacion = mock_class  

    with patch('importlib.import_module', return_value=mock_module):
        
        excel_instance._define_what_process_run(record, fic_id)

        mock_class_instance.process.assert_called_once()



from openpyxl import load_workbook
import importlib
import pdb
import time
from bd.db.querys_db import QuerysDB
import bd.db.queries as queries
from bd.db.insert_into_histories import InsertIntoHistories
from .shared import get_class_name


class ExcelProcess:
    ENTITY = {
        2: "ficha_caracterizacion",
        4: "registro_academico",
        6: "instructor_x_ficha",
    }

    def __init__(self, oc_connection, pg_connection):
        self.oc_connection = oc_connection
        self.pg_connection = pg_connection

    def execute(self):
        try:

            workbook = load_workbook(
                filename="attachments/Grupos_fichas_desde_01072020.xlsx"
            )
            sheet = workbook.active
            valores_columna = [
                row[0]
                for row in sheet.iter_rows(
                    min_row=1, max_col=1, max_row=sheet.max_row, values_only=True
                )
            ]
            #print(len(valores_columna))
            for valor in valores_columna[1:]:
                #print(valor)
                self._define_what_process_run(None, valor)

        except Exception as e:
            print("excepcionx", e)
            insert_into_histories = InsertIntoHistories(self.pg_connection)
            insert_into_histories.insert(
                (
                    "No se encontraron Fichas para procesar en la tabla INDICE_CAMBIO ",
                    "sin datos",
                    "sin datos",
                )
            )

    def _define_what_process_run(self, record, fic_id):
        # print(f'40 {record}')
        module_name = f"entity.{self.ENTITY[2]}"
        module = importlib.import_module(module_name)
        class_name = get_class_name(module_name.replace("_", " "))
        import_class = getattr(module, class_name)
        instance_class = import_class(
            self.oc_connection, self.pg_connection, record, fic_id
        )
        instance_class.process()


